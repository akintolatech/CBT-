from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import cache_control
from django.views.decorators.cache import never_cache
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render, redirect
from .forms import NameForm
from .models import CustomUser, Question, Test, TestResult, Class
from django.views.decorators.csrf import csrf_protect, csrf_exempt
import datetime
from django.utils import timezone
from django.contrib import messages

import openpyxl
from django.http import HttpResponse

now = timezone.now()

from django.contrib.sessions.models import Session
from django.contrib.auth import get_user_model
from django.db.models import Q


# Create your views here.
def register(request):
    if request.method == 'POST':
        form = NameForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('login')
    else:
        form = NameForm()

    context = {
        'error': 'Something went wrong',
        'form': form
    }

    return render(request, 'authenticator/register.html', context)


# @login_required
User = get_user_model()


def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']

        user = authenticate(request, username=username, password=password)

        if user is not None:
            # Check if the user has an active session on another device
            active_sessions = Session.objects.filter(Q(expire_date__gte=timezone.now()) | Q(expire_date=None))
            for session in active_sessions:
                session_data = session.get_decoded()
                session_user_id = session_data.get('_auth_user_id')
                if session_user_id is not None and user.id == int(session_user_id):
                    # Another active session found, prevent login
                    context = {'error': 'You are already logged in on another device.'}
                    return render(request, 'authenticator/login.html', context)

            # Log in the user and create a new session
            login(request, user)
            return redirect('dashboard')
        else:
            context = {'error': 'Invalid username or password'}
            return render(request, 'authenticator/login.html', context)
    else:
        return render(request, 'authenticator/login.html')


def index(request):
    return render(request, 'authenticator/index.html')


# result viewer login
def rv_login(request):
    if request.method == 'POST':

        username = request.POST['username']
        password = request.POST['password']

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect('results_viewer')
        else:
            context = {'error': 'Invalid username or password'}
            return render(request, 'authenticator/rvlogin.html', context)
    else:
        return render(request, 'authenticator/rvlogin.html')


# v 0.6
@login_required(login_url='/login/')
def dashboard(request):
    if request.user.is_authenticated:
        # render the test page
        # tests = Test.objects.all()

        classes = Class.objects.all()
        selected_class = request.user.classes

        # display the test based on science or art student
        if selected_class == "SCIENCE":
            tests = Test.objects.filter(classes__id=1).order_by('name')
        else:
            tests = Test.objects.filter(classes__id=2).order_by('name')

        written_tests = Test.objects.filter(written_by=request.user)

        context = {
            'tests': tests,
            'written_tests': written_tests,
            'name': request.user.username,
        }

        return render(request, 'authenticator/dashboard.html', context)
    else:
        return redirect('login')


def test_questions(request, test_id):
    if request.user.is_authenticated:
        tests = Test.objects.get(id=test_id)
        tests_question = get_object_or_404(Test, pk=test_id)

        context = {
            'tests': tests,
            'test': tests_question,
            'name': request.user.username,
        }

        return render(request, 'authenticator/test.html', context)

    else:
        return redirect('login')


@csrf_exempt
def mark_test(request, test_id):
    if request.user.is_authenticated:

        tests_question = get_object_or_404(Test, pk=test_id)

        # Check if the user has already written the test
        # if tests_question.written_by.filter(pk=request.user.pk).exists():
        #     return redirect('results')  # Redirect to results page or any other appropriate action

        questions = Question.objects.all()
        results = TestResult
        score = 0

        for question in questions:
            if request.POST.get(str(question.id)) == question.correct_option:
                score += 1

        results.objects.create(
            test=tests_question,
            user=request.user,
            username=request.user.username,
            svc_no=request.user.password,
            score=score,
            classes=request.user.classes,
            desc=tests_question.name,
            date=now,
        )

        tests_question.written_by.add(request.user)  # Add the user to the written_by ManyToMany field
        tests_question.save()
        return redirect('results')

    else:
        return redirect('login')


def view_results(request):
    results = TestResult.objects.filter(user=request.user)
    context = {

        'results': results,

    }
    return render(request, 'authenticator/results.html', context)


@staff_member_required
@login_required(login_url='index')  # Redirects non-authenticated users to the login page
def results_viewer(request):
    results = TestResult.objects.all()

    if request.method == 'POST':

        # Extract variables to be used for querying the db
        search_name = request.POST.get('search_name')
        selected_class = request.POST.get('selected_class')
        selected_subject = request.POST.get('selected_subject')

        # Perform filtering based on the selected criteria
        # Assuming 'results' is a queryset of the queried results
        filtered_results = TestResult.objects.all()

        if search_name:
            filtered_results = results.filter(username__icontains=search_name)

        if selected_class != 'All' and selected_subject != 'All':
            filtered_results = results.filter(classes=selected_class, desc=selected_subject)

        if 'export' in request.POST:  # Check if the export button is clicked
            if filtered_results.exists():
                # Export the filtered results to Excel
                wb = openpyxl.Workbook()
                ws = wb.active

                # Write column headers
                ws['A1'] = 'User'
                ws['B1'] = 'Test'
                ws['C1'] = 'Classes'
                ws['D1'] = 'Score'

                # Write filtered results to rows
                row_num = 2
                for result in filtered_results:
                    ws.cell(row=row_num, column=1, value=result.username)
                    ws.cell(row=row_num, column=2, value=result.desc)
                    ws.cell(row=row_num, column=3, value=result.classes)
                    ws.cell(row=row_num, column=4, value=result.score)
                    row_num += 1

                # Set response headers for Excel file download
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=exported_results.xlsx'

                # Save the workbook to the response
                wb.save(response)

                return response

        # Return the filtered results to the template or No results found
        if filtered_results.exists():
            context = {
                'results': filtered_results,
            }
        else:
            context = {
                'results': None,
                'not_found_message': 'No results found.',
            }

            return render(request, 'authenticator/results_viewer.html', context)

        return render(request, 'authenticator/results_viewer.html', context)

    else:
        context = {

            'results': results,

        }
        return render(request, 'authenticator/results_viewer.html', context)


def logout_view(request):
    logout(request)
    return redirect('index')
